import time

from validate_email import validate_email
import openpyxl
import DNS

DNS.defaults['server'] = ['8.8.8.8', '8.8.4.4']


class EmailValidator:

    def __init__(self, file_name):
        self.file_name = file_name

    def get_info(self):
        book = openpyxl.open(self.file_name, read_only=True)

        sheet = book.active

        emails = []

        for row in range(1, sheet.max_row + 1):
            email = sheet[row][0].value

            emails.append(email)

        return emails

    @staticmethod
    def check_address(check_list):

        check_result = []
        counter = 0

        for address in check_list:
            try:
                check_syntax = validate_email(address)
                check_existence = validate_email(address, verify=True)
                check_dns = validate_email(address, check_mx=True)

                if check_syntax and check_existence and check_dns:
                    check_result.append('Валидный')
                else:
                    check_result.append('Не валидный')

                if counter % 100 == 0:
                    time.sleep(5)
                    print(f'Проверено {counter} адресов')
                    counter += 1
                else:
                    counter += 1
                    continue

            except TimeoutError:
                check_result.append('Ошибка валидации')

        return check_result

    def write_data(self, data):
        book = openpyxl.load_workbook(self.file_name)

        sheet = book.active

        row = 1

        for record in data:

            sheet.cell(row=row, column=2, value=record)

            row += 1

        book.save(self.file_name)

        return True


if __name__ == '__main__':
    test = EmailValidator('Файл проверки.xlsx')
    start_time = time.time()
    addresses = test.get_info()
    check_result = test.check_address(addresses)
    flag = test.write_data(check_result)
    print("--- %s seconds ---" % (time.time() - start_time))

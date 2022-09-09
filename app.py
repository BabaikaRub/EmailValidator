import traceback
import sys
import time

from PyQt5 import QtWidgets
from PyQt5.QtCore import QObject, pyqtSignal, QThread
from PyQt5.QtWidgets import QMessageBox, QFileDialog
from openpyxl.utils.exceptions import InvalidFileException
from validate_email import validate_email
import openpyxl
import DNS

from ui import Ui_MainWindow


DNS.defaults['server'] = ['8.8.8.8', '8.8.4.4']


class App(QtWidgets.QMainWindow):

    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.init_ui()

        self.ui.pushButton.clicked.connect(self.load_file)

    def load_file(self):
        dialog = QFileDialog.getOpenFileName(self, "Select Excel file to import", "", "Excel (*.xls *.xlsx)")
        file_name = dialog[0]

        self.start_validate(file_name)

    @staticmethod
    def popup_info():
        msg = QMessageBox()
        msg.setWindowTitle('Информация')
        msg.setText('Для начала работы выберите файл!')

        x = msg.exec_()

    def show_end(self):
        self.ui.label_2.setText('Файл записан')
        self.ui.pushButton.setEnabled(1)

    def show_info(self, info):
        self.ui.label_2.setText(f'Проверено {info} адресов')

    def start_info(self):
        self.ui.label_2.setText('В процессе...')

    def init_ui(self):
        self.setWindowTitle('EmailValidator')

    def start_validate(self, file_name):
        self.obj = Validator(file_name)
        self.t = QThread()
        self.obj.moveToThread(self.t)
        self.t.started.connect(self.obj.start)
        self.obj.startSignal.connect(self.start_info)
        self.obj.updateSignal.connect(self.show_info)
        self.obj.finishSignal.connect(self.t.quit)
        self.obj.finishSignal.connect(self.show_end)

        self.t.start()

        self.ui.pushButton.setEnabled(0)


class Validator(QObject):
    finishSignal = pyqtSignal()
    updateSignal = pyqtSignal(int)
    startSignal = pyqtSignal()

    def __init__(self, file_name):
        super().__init__()
        self.file_name = file_name

    def start(self):
        try:
            excel_writer = ExcelWorker(self.file_name)
            self.startSignal.emit()
            addresses = excel_writer.get_info()
            check_result = self.check_address(addresses)
            excel_writer.write_data(check_result)

            self.finishSignal.emit()

        except InvalidFileException:
            App.popup_info()

    def check_address(self, check_list):

        check_result = []
        counter = 1

        for address in check_list:
            try:
                try:
                    check_syntax = validate_email(address)
                    check_existence = validate_email(address, verify=True)
                    check_dns = validate_email(address, check_mx=True)

                    if check_syntax and check_existence and check_dns:
                        check_result.append('Валидный')
                    else:
                        check_result.append('Не валидный')

                except UnicodeEncodeError:
                    check_result.append('Не валидный')

                if counter % 100 == 0:
                    time.sleep(5)
                    self.updateSignal.emit(counter)
                    counter += 1
                else:
                    counter += 1
                    continue

            except TimeoutError:
                check_result.append('Ошибка валидации')

        return check_result


class ExcelWorker:

    def __init__(self, file_name):
        self.file_name = file_name

    def get_info(self):
        book = openpyxl.open(self.file_name, read_only=True)

        sheet = book.active

        emails = []

        for row in range(1, sheet.max_row + 1):
            email = sheet[row][0].value

            emails.append(email)

        return emails

    def write_data(self, data):
        book = openpyxl.load_workbook(self.file_name)

        sheet = book.active

        row = 1

        for record in data:

            sheet.cell(row=row, column=2, value=record)

            row += 1

        book.save(self.file_name)


def main():

    def log_uncaught_exceptions(ex_cls, ex, tb):
        text = '{}: {}:\n'.format(ex_cls.__name__, ex)
        text += ''.join(traceback.format_tb(tb))

        print(text)
        QtWidgets.QMessageBox.critical(None, 'Error', text)
        quit()

    sys.excepthook = log_uncaught_exceptions

    app = QtWidgets.QApplication(sys.argv)
    application = App()
    application.show()

    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
